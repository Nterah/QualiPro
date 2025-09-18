# app/pqp/ingest/ai_import.py
from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any, Dict, List, Tuple, Optional

from openpyxl import load_workbook

# Pure metadata – safe to import (no app/route side effects)
from app.pqp.sections import SECTION_DEFS, DEFAULT_SECTION_TITLES
# Model only (no routes) – avoids circular imports
from app.pqp.pqp_models import PQPSection


# -------------------------- helpers --------------------------

def _clean_cell(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()

def _iso_date_like(v):
    s = _clean_cell(v)
    if not s:
        return ""
    # already ISO
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    # dd/mm/yyyy, dd-mm-yyyy, dd.mm.yyyy
    m = re.fullmatch(r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{4})", s)
    if m:
        d, mth, y = m.groups()
        return f"{int(y):04d}-{int(mth):02d}-{int(d):02d}"
    # last resort: leave as-is
    return s

def _load_section_rows(sec: PQPSection) -> List[dict]:
    payload = None
    # Support either rows_json or content (Text)
    if hasattr(sec, "rows_json") and getattr(sec, "rows_json", None):
        try:
            payload = json.loads(sec.rows_json)
        except Exception:
            payload = None
    if payload is None and getattr(sec, "content", None):
        try:
            payload = json.loads(sec.content)
        except Exception:
            payload = None
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and isinstance(payload.get("rows"), list):
        return payload["rows"]
    return []

def _dump_section_rows(rows: List[dict]) -> str:
    return json.dumps(rows, ensure_ascii=False)


# -------------------------- parsing --------------------------

def parse_workbook_to_payload(stream, project_code: Optional[str] = None
                             ) -> Tuple[Dict[str, Any], List[str], Optional[str]]:
    """
    Returns (payload, issues, detected_code).
    payload = {"code": <str>, "sections": [{"index": n, "rows": [...]}, ...]}
    """
    issues: List[str] = []
    wb = load_workbook(filename=BytesIO(stream.read()), data_only=True)

    # --- naive code detection from first couple of sheets ---
    detected_code: Optional[str] = None
    try:
        for ws in wb.worksheets[:2]:
            for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True):
                for val in row:
                    s = _clean_cell(val)
                    if not s:
                        continue
                    # heuristic: letters + 2+ digits (eg "P700", "GENL-PQP-04")
                    if re.match(r"^[A-Za-z]*\d{2,}[A-Za-z0-9\-]*$", s):
                        detected_code = s
                        break
                if detected_code:
                    break
            if detected_code:
                break
    except Exception as e:
        issues.append(f"Code detection failed: {e}")

    code_for_payload = project_code or detected_code or ""
    payload: Dict[str, Any] = {"code": code_for_payload, "sections": []}

    # --- map first 9 worksheets to sections 1..9 ---
    sheets = wb.worksheets[:9]
    for idx, ws in enumerate(sheets, start=1):
        expected_cols = SECTION_DEFS[idx - 1]

        # header = row 1
        header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header = [(_clean_cell(h) or "") for h in header_cells]
        if not any(header):
            issues.append(f"Sheet '{ws.title}': empty header; skipping")
            payload["sections"].append({"index": idx, "rows": []})
            continue

        rows: List[Dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rec: Dict[str, Any] = {}
            any_val = False
            for j, col in enumerate(header):
                if col in expected_cols:
                    val = _clean_cell(r[j] if j < len(r) else "")
                    if "date" in col.lower():
                        val = _iso_date_like(val)
                    rec[col] = val
                    if val not in ("", None):
                        any_val = True
            if any_val:
                if "id" in expected_cols and "id" not in rec:
                    rec["id"] = ""
                rows.append(rec)

        payload["sections"].append({"index": idx, "rows": rows})

    return payload, issues, detected_code


# -------------------------- commit --------------------------

def commit_payload(payload: Dict[str, Any],
                   project_code: Optional[str],
                   db_session) -> Tuple[bool, List[str]]:
    """
    Write parsed payload into PQPSection per section.
    Merge by 'id' when present, otherwise append.
    Returns (ok, issues).
    """
    issues: List[str] = []

    code = (project_code or payload.get("code") or "").strip()
    if not code:
        return False, ["No project code provided or detected"]

    sections = payload.get("sections", [])
    if not isinstance(sections, list):
        return False, ["Invalid 'sections' structure"]

    total_created = 0
    total_updated = 0

    for sec_payload in sections:
        try:
            idx = int(sec_payload.get("index") or 0)
        except Exception:
            idx = 0
        if idx < 1 or idx > len(SECTION_DEFS):
            issues.append(f"Skipping invalid section index: {sec_payload.get('index')!r}")
            continue

        new_rows = sec_payload.get("rows") or []
        if not isinstance(new_rows, list):
            issues.append(f"Section {idx}: rows not a list; skipping")
            continue

        # fetch/create PQPSection
        sec = PQPSection.query.filter_by(project_code=code, section_number=idx).first()
        if not sec:
            sec = PQPSection(
                project_code=code,
                section_number=idx,
                title=DEFAULT_SECTION_TITLES.get(idx, f"Section {idx}")
            )
            db_session.add(sec)
            db_session.flush()

        existing = _load_section_rows(sec)

        # merge by id if present
        # build index for existing rows with id
        existing_by_id = {}
        for er in existing:
            if isinstance(er, dict):
                rid = str(er.get("id") or "").strip()
                if rid:
                    existing_by_id[rid] = er

        for nr in new_rows:
            if not isinstance(nr, dict):
                continue
            rid = str(nr.get("id") or "").strip()
            if rid and rid in existing_by_id:
                existing_by_id[rid].update(nr)
                total_updated += 1
            else:
                # ensure id key exists if schema includes it
                if "id" in SECTION_DEFS[idx - 1] and "id" not in nr:
                    nr["id"] = ""
                existing.append(nr)
                total_created += 1

        # write back
        if hasattr(sec, "rows_json"):
            sec.rows_json = _dump_section_rows(existing)
        else:
            sec.content = _dump_section_rows(existing)

    db_session.commit()
    if total_created or total_updated:
        issues.append(f"Upserted rows — created: {total_created}, updated: {total_updated}")
    return True, issues
